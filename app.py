import streamlit as st
import json
from datetime import datetime
import pandas as pd
from io import BytesIO
import base64

# 페이지 설정
st.set_page_config(
    page_title="위험성평가 작성 프로그램",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# 커스텀 CSS - 나눔고딕 폰트 및 스타일링
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Nanum+Gothic:wght@400;700;800&display=swap');
    
    * {
        font-family: 'Nanum Gothic', sans-serif !important;
    }
    
    .main {
        background-color: #f0f4f8;
    }
    
    .stButton > button {
        background: linear-gradient(90deg, #3b82f6 0%, #2563eb 100%);
        color: white;
        border: none;
        padding: 0.5rem 2rem;
        border-radius: 20px;
        font-weight: bold;
        transition: all 0.3s;
    }
    
    .stButton > button:hover {
        transform: scale(1.05);
        box-shadow: 0 5px 15px rgba(37, 99, 235, 0.4);
    }
    
    .cover-container {
        background: white;
        border: 2px solid #1f2937;
        border-radius: 10px;
        padding: 40px;
        max-width: 1000px;
        margin: 0 auto;
        box-shadow: 0 10px 30px rgba(0, 0, 0, 0.1);
    }
    
    .title-gradient {
        background: linear-gradient(90deg, #3b82f6 0%, #6366f1 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        font-size: 2.5rem;
        font-weight: 800;
        text-align: center;
        margin: 20px 0;
    }
    
    .approval-table {
        border-collapse: collapse;
        margin: 30px auto;
    }
    
    .approval-table td {
        border: 1px solid #d1d5db;
        padding: 10px;
        text-align: center;
    }
    
    .approval-header {
        background: linear-gradient(90deg, #fde047 0%, #facc15 100%);
        font-weight: bold;
        width: 60px;
    }
    
    input[type="text"] {
        border: 2px solid #e5e7eb;
        border-radius: 8px;
        padding: 8px 12px;
        width: 100%;
        transition: all 0.3s;
    }
    
    input[type="text"]:focus {
        border-color: #3b82f6;
        outline: none;
        box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.1);
    }
    
    /* 유해위험요인 테이블 스타일 - 줄간격 조정 */
    .hazard-table {
        margin: 0 !important;
        padding: 0 !important;
    }
    
    .hazard-table .stCheckbox {
        margin-bottom: -20px !important;
        padding: 0px !important;
    }
    
    .hazard-table label {
        margin-bottom: 0 !important;
        font-size: 12px !important;
        padding: 0px !important;
        line-height: 1.2 !important;
    }
    
    /* 체크박스 간격 조정 */
    .stCheckbox > div {
        margin-bottom: -15px !important;
        padding: 0 !important;
    }
    
    .stCheckbox > label > div {
        padding: 0 !important;
    }
    
    /* 체크박스 컨테이너 간격 제거 */
    [data-testid="stVerticalBlock"] > div:has(.stCheckbox) {
        gap: 0 !important;
    }
    
    /* 체크박스 자체 크기 조정 */
    .stCheckbox input[type="checkbox"] {
        margin-right: 5px !important;
    }
</style>
""", unsafe_allow_html=True)

# 세션 상태 초기화
if 'form_data' not in st.session_state:
    st.session_state.form_data = {
        'year': '',
        'company_name': '',
        'address': '',
        'phone': '',
        'fax': '',
        'approvers': [
            {'position': '', 'name': ''},
            {'position': '', 'name': ''},
            {'position': '', 'name': ''},
            {'position': '', 'name': ''}
        ]
    }

# 제목
st.markdown('<h1 style="text-align: center; color: #1f2937;">위험성평가 작성 프로그램</h1>', unsafe_allow_html=True)
st.markdown('---')

# 탭 생성
tab1, tab2, tab3, tab4 = st.tabs(["📄 표지", "📊 사업장 개요", "⚠️ 위험정보", "📋 유해위험요인"])

with tab1:
    st.markdown('<div class="cover-container">', unsafe_allow_html=True)
    
    # 연도 입력
    col1, col2, col3 = st.columns([2, 1, 2])
    with col2:
        year = st.text_input(
            "연도",
            value=st.session_state.form_data['year'],
            max_chars=2,
            placeholder="24",
            label_visibility="collapsed"
        )
        if year:
            st.session_state.form_data['year'] = year
        
        st.markdown(f'<p style="text-align: center; font-size: 1.5rem; font-weight: bold;">20{year if year else "( )"}년도</p>', unsafe_allow_html=True)
    
    # 제목
    st.markdown('<h1 class="title-gradient">위험성평가 결과서</h1>', unsafe_allow_html=True)
    
    # 결재란
    st.markdown('<h3 style="text-align: center; margin-top: 30px;">결재란</h3>', unsafe_allow_html=True)
    
    col1, col2, col3, col4, col5 = st.columns(5)
    
    # 결재 헤더
    with col1:
        st.markdown('<div style="border: 1px solid #d1d5db; background: linear-gradient(90deg, #fde047 0%, #facc15 100%); padding: 20px 10px; text-align: center; font-weight: bold;">결재</div>', unsafe_allow_html=True)
    
    # 결재자 입력
    for i, col in enumerate([col2, col3, col4, col5]):
        with col:
            position = st.text_input(
                f"직위{i+1}",
                value=st.session_state.form_data['approvers'][i]['position'],
                placeholder="직위",
                key=f"position_{i}",
                label_visibility="collapsed"
            )
            name = st.text_input(
                f"성명{i+1}",
                value=st.session_state.form_data['approvers'][i]['name'],
                placeholder="성명",
                key=f"name_{i}",
                label_visibility="collapsed"
            )
            
            st.session_state.form_data['approvers'][i]['position'] = position
            st.session_state.form_data['approvers'][i]['name'] = name
    
    # 회사 정보
    st.markdown('<h3 style="text-align: center; margin-top: 40px;">회사 정보</h3>', unsafe_allow_html=True)
    
    company_name = st.text_input(
        "회사명",
        value=st.session_state.form_data['company_name'],
        placeholder="회사명을 입력하세요",
        label_visibility="collapsed"
    )
    st.session_state.form_data['company_name'] = company_name
    
    address = st.text_input(
        "주소",
        value=st.session_state.form_data['address'],
        placeholder="주소를 입력하세요",
        label_visibility="collapsed"
    )
    st.session_state.form_data['address'] = address
    
    col1, col2 = st.columns(2)
    with col1:
        phone = st.text_input(
            "전화번호",
            value=st.session_state.form_data['phone'],
            placeholder="전화번호",
            label_visibility="collapsed"
        )
        st.session_state.form_data['phone'] = phone
    
    with col2:
        fax = st.text_input(
            "팩스번호",
            value=st.session_state.form_data['fax'],
            placeholder="팩스번호",
            label_visibility="collapsed"
        )
        st.session_state.form_data['fax'] = fax
    
    st.markdown('</div>', unsafe_allow_html=True)
    
    # 버튼들
    st.markdown('<br><br>', unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 1, 1])
    
    with col2:
        if st.button("💾 표지 엑셀 저장", use_container_width=True):
            # 엑셀로 저장
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # 빈 데이터프레임 생성 (시트 생성용)
                df = pd.DataFrame()
                df.to_excel(writer, sheet_name='표지', index=False)
                
                # 워크시트 가져오기
                workbook = writer.book
                worksheet = writer.sheets['표지']
                
                # 서식 설정
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
                
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                center_alignment = Alignment(horizontal='center', vertical='center')
                
                # 연도 (병합 셀)
                worksheet.merge_cells('C2:E3')
                year_cell = worksheet['C2']
                year_cell.value = f"20{st.session_state.form_data['year']}년도"
                year_cell.font = Font(size=20, bold=True)
                year_cell.alignment = center_alignment
                
                # 제목 (병합 셀)
                worksheet.merge_cells('B5:F7')
                title_cell = worksheet['B5']
                title_cell.value = "위험성평가 결과서"
                title_cell.font = Font(size=28, bold=True)
                title_cell.alignment = center_alignment
                
                # 결재란 헤더
                worksheet.merge_cells('B10:B12')
                approval_header = worksheet['B10']
                approval_header.value = "결재"
                approval_header.fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
                approval_header.font = Font(bold=True)
                approval_header.alignment = center_alignment
                approval_header.border = thin_border
                
                # 결재자 정보
                columns = ['C', 'D', 'E', 'F']
                for i, col in enumerate(columns):
                    # 직위
                    position_cell = worksheet[f'{col}10']
                    position_cell.value = st.session_state.form_data['approvers'][i]['position']
                    position_cell.alignment = center_alignment
                    position_cell.border = thin_border
                    
                    # 서명 공간
                    worksheet.merge_cells(f'{col}11:{col}11')
                    sign_cell = worksheet[f'{col}11']
                    sign_cell.border = thin_border
                    worksheet.row_dimensions[11].height = 40
                    
                    # 성명
                    name_cell = worksheet[f'{col}12']
                    name_cell.value = st.session_state.form_data['approvers'][i]['name']
                    name_cell.alignment = center_alignment
                    name_cell.border = thin_border
                
                # 회사 정보 (하단)
                info_start_row = 15
                worksheet[f'C{info_start_row}'].value = st.session_state.form_data['company_name']
                worksheet[f'C{info_start_row}'].font = Font(size=16, bold=True)
                worksheet[f'C{info_start_row}'].alignment = center_alignment
                
                worksheet[f'C{info_start_row+2}'].value = f"주소: {st.session_state.form_data['address']}"
                worksheet[f'C{info_start_row+3}'].value = f"전화: {st.session_state.form_data['phone']}"
                worksheet[f'C{info_start_row+4}'].value = f"팩스: {st.session_state.form_data['fax']}"
                
                # 열 너비 조정
                worksheet.column_dimensions['A'].width = 5
                worksheet.column_dimensions['B'].width = 15
                worksheet.column_dimensions['C'].width = 15
                worksheet.column_dimensions['D'].width = 15
                worksheet.column_dimensions['E'].width = 15
                worksheet.column_dimensions['F'].width = 15
                worksheet.column_dimensions['G'].width = 5
            
            output.seek(0)
            b64 = base64.b64encode(output.read()).decode()
            href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="위험성평가_표지_{st.session_state.form_data.get("year", "YYYY")}.xlsx">📥 엑셀 파일 다운로드</a>'
            st.markdown(href, unsafe_allow_html=True)
            st.success("표지가 엑셀 파일로 저장되었습니다!")

with tab2:
    st.markdown('<h2 style="text-align: center; color: #1f2937;">1. 사업장 개요</h2>', unsafe_allow_html=True)
    
    # 세션 상태 초기화 - 사업장 개요
    if 'business_info' not in st.session_state:
        st.session_state.business_info = {
            'business_name': '',
            'main_product': '',
            'evaluation_date': '',
            'representative': '',
            'employee_count': '',
            'evaluator': ''
        }
    
    # 사업장 정보 입력
    col1, col2 = st.columns(2)
    
    with col1:
        st.session_state.business_info['business_name'] = st.text_input(
            "사업장명",
            value=st.session_state.business_info['business_name'],
            placeholder="사업장명을 입력하세요"
        )
        st.session_state.business_info['main_product'] = st.text_input(
            "주요생산품",
            value=st.session_state.business_info['main_product'],
            placeholder="주요생산품을 입력하세요"
        )
        st.session_state.business_info['evaluation_date'] = st.date_input(
            "평가일자",
            value=None,
            format="YYYY/MM/DD"
        )
    
    with col2:
        st.session_state.business_info['representative'] = st.text_input(
            "대표자",
            value=st.session_state.business_info['representative'],
            placeholder="대표자명을 입력하세요"
        )
        st.session_state.business_info['employee_count'] = st.text_input(
            "근로자수",
            value=st.session_state.business_info['employee_count'],
            placeholder="근로자수를 입력하세요"
        )
        st.session_state.business_info['evaluator'] = st.text_input(
            "평가자",
            value=st.session_state.business_info['evaluator'],
            placeholder="평가자명을 입력하세요"
        )
    
    st.markdown('<hr>', unsafe_allow_html=True)
    st.markdown('<h3 style="text-align: center; color: #1f2937;">공정도</h3>', unsafe_allow_html=True)
    
    # 공정 수 관리
    if 'process_count' not in st.session_state:
        st.session_state.process_count = 5
    
    if 'processes' not in st.session_state:
        st.session_state.processes = [
            {
                'name': '',
                'photo': None,
                'description': '',
                'equipment': '',
                'hazardous_material': '',
                'hazardous_factor': ''
            } for _ in range(5)
        ]
    
    # 공정 추가/삭제 버튼
    col1, col2, col3 = st.columns([1, 2, 1])
    with col1:
        if st.button("➕ 공정 추가"):
            st.session_state.process_count += 1
            st.session_state.processes.append({
                'name': '',
                'photo': None,
                'description': '',
                'equipment': '',
                'hazardous_material': '',
                'hazardous_factor': ''
            })
            st.rerun()
    
    with col3:
        if st.button("➖ 공정 삭제") and st.session_state.process_count > 1:
            st.session_state.process_count -= 1
            st.session_state.processes.pop()
            st.rerun()
    
    # 공정을 5개씩 그룹으로 나누어 표시
    process_groups = []
    for i in range(0, st.session_state.process_count, 5):
        process_groups.append(range(i, min(i + 5, st.session_state.process_count)))
    
    # 각 그룹별로 공정 표시
    for group_idx, process_group in enumerate(process_groups):
        if group_idx > 0:
            st.markdown('<hr style="margin: 30px 0;">', unsafe_allow_html=True)
            
        # 이 그룹의 공정 수 (최대 5개)
        group_size = len(process_group)
        
        # 컬럼 생성 (공정들만)
        cols = st.columns(group_size)
        
        # 각 공정별 입력 필드
        for col_idx, process_idx in enumerate(process_group):
            with cols[col_idx]:
                # 공정명
                st.markdown('<div style="font-weight: bold; margin-bottom: 5px;">공정명</div>', unsafe_allow_html=True)
                st.session_state.processes[process_idx]['name'] = st.text_input(
                    f"공정명 {process_idx+1}",
                    value=st.session_state.processes[process_idx]['name'],
                    placeholder=f"공정 {process_idx+1}",
                    key=f"process_name_{process_idx}",
                    label_visibility="collapsed"
                )
                
                # 화살표 표시 (각 그룹의 첫 번째 공정 제외)
                if col_idx > 0 or (group_idx > 0 and col_idx == 0):
                    st.markdown('<div style="text-align: center; font-size: 20px; color: #6b7280; margin: 5px 0;">→</div>', unsafe_allow_html=True)
                else:
                    st.markdown('<div style="margin: 5px 0; height: 28px;"></div>', unsafe_allow_html=True)
                
                # 공정사진
                st.markdown('<div style="font-weight: bold; margin-bottom: 5px;">공정사진</div>', unsafe_allow_html=True)
                photo = st.file_uploader(
                    f"공정사진 {process_idx+1}",
                    type=['png', 'jpg', 'jpeg'],
                    key=f"process_photo_{process_idx}",
                    label_visibility="collapsed"
                )
                if photo:
                    st.session_state.processes[process_idx]['photo'] = photo
                    st.image(photo, use_column_width=True)
                else:
                    st.markdown('<div style="height: 120px; border: 2px dashed #d1d5db; display: flex; align-items: center; justify-content: center; color: #9ca3af; background-color: #f9fafb;">사진 업로드<br>클릭 또는 드래그</div>', unsafe_allow_html=True)
                
                # 간격
                st.markdown('<div style="margin: 10px 0;"></div>', unsafe_allow_html=True)
                
                # 공정설명
                st.markdown('<div style="font-weight: bold; margin-bottom: 5px;">공정설명</div>', unsafe_allow_html=True)
                st.session_state.processes[process_idx]['description'] = st.text_area(
                    f"공정설명 {process_idx+1}",
                    value=st.session_state.processes[process_idx]['description'],
                    placeholder="공정 설명",
                    key=f"process_desc_{process_idx}",
                    height=100,
                    label_visibility="collapsed"
                )
                
                # 간격
                st.markdown('<div style="margin: 10px 0;"></div>', unsafe_allow_html=True)
                
                # 주요기계기구
                st.markdown('<div style="font-weight: bold; margin-bottom: 5px;">주요기계기구</div>', unsafe_allow_html=True)
                st.session_state.processes[process_idx]['equipment'] = st.text_area(
                    f"주요기계기구 {process_idx+1}",
                    value=st.session_state.processes[process_idx]['equipment'],
                    placeholder="주요기계기구",
                    key=f"process_equip_{process_idx}",
                    height=100,
                    label_visibility="collapsed"
                )
                
                # 간격
                st.markdown('<div style="margin: 10px 0;"></div>', unsafe_allow_html=True)
                
                # 유해위험물질
                st.markdown('<div style="font-weight: bold; margin-bottom: 5px;">유해위험물질</div>', unsafe_allow_html=True)
                st.session_state.processes[process_idx]['hazardous_material'] = st.text_area(
                    f"유해위험물질 {process_idx+1}",
                    value=st.session_state.processes[process_idx]['hazardous_material'],
                    placeholder="유해위험물질",
                    key=f"process_material_{process_idx}",
                    height=100,
                    label_visibility="collapsed"
                )
                
                # 간격
                st.markdown('<div style="margin: 10px 0;"></div>', unsafe_allow_html=True)
                
                # 유해위험요인
                st.markdown('<div style="font-weight: bold; margin-bottom: 5px;">유해위험요인</div>', unsafe_allow_html=True)
                st.session_state.processes[process_idx]['hazardous_factor'] = st.text_area(
                    f"유해위험요인 {process_idx+1}",
                    value=st.session_state.processes[process_idx]['hazardous_factor'],
                    placeholder="유해위험요인",
                    key=f"process_factor_{process_idx}",
                    height=100,
                    label_visibility="collapsed"
                )
    
    # 데이터 저장 버튼
    st.markdown('<br>', unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 1, 1])
    with col2:
        if st.button("💾 사업장 개요 엑셀 저장", use_container_width=True, key="save_tab2"):
            # 사업장 개요 데이터
            overview_df = pd.DataFrame([st.session_state.business_info])
            
            # 공정 데이터
            process_list = []
            for process in st.session_state.processes:
                if process['name']:
                    process_list.append({
                        '공정명': process['name'],
                        '공정설명': process['description'],
                        '주요기계기구': process['equipment'],
                        '유해위험물질': process['hazardous_material'],
                        '유해위험요인': process['hazardous_factor']
                    })
            
            process_df = pd.DataFrame(process_list)
            
            # 엑셀로 저장
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                overview_df.to_excel(writer, sheet_name='사업장정보', index=False)
                if not process_df.empty:
                    process_df.to_excel(writer, sheet_name='공정정보', index=False)
                
                # 서식 설정
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                
                header_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
                header_font = Font(bold=True, size=12)
                header_alignment = Alignment(horizontal='center', vertical='center')
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # 각 시트에 서식 적용
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    
                    # 헤더 서식
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = header_alignment
                        cell.border = thin_border
                    
                    # 열 너비 자동 조정
                    for column in worksheet.columns:
                        max_length = 0
                        column = [cell for cell in column]
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2) * 1.2
                        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
            output.seek(0)
            b64 = base64.b64encode(output.read()).decode()
            href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="위험성평가_사업장개요_{datetime.now().strftime("%Y%m%d")}.xlsx">📥 엑셀 파일 다운로드</a>'
            st.markdown(href, unsafe_allow_html=True)
            st.success("사업장 개요가 엑셀 파일로 저장되었습니다!")

with tab3:
    st.markdown('<h2 style="text-align: center; color: #1f2937;">안전보건상 위험정보</h2>', unsafe_allow_html=True)
    
    # 상단 정보 테이블 스타일
    st.markdown("""
    <style>
    .info-header {
        background-color: #fef3c7;
        border: 1px solid #d97706;
        padding: 10px;
        font-weight: bold;
        text-align: center;
        min-width: 120px;
    }
    .process-table {
        width: 100%;
        border-collapse: collapse;
        margin-top: 20px;
    }
    .process-table th, .process-table td {
        border: 1px solid #d97706;
        padding: 10px;
        text-align: center;
    }
    .process-header {
        background-color: #fef3c7;
        font-weight: bold;
        font-size: 16px;
    }
    .sub-header {
        background-color: #fef3c7;
        font-size: 14px;
        font-weight: normal;
    }
    .stTextInput input, .stTextArea textarea, .stSelectbox select {
        font-size: 16px !important;
        padding: 10px !important;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # 상단 정보 입력
    col1, col2, col3, col4 = st.columns([1, 2, 1, 2])
    with col1:
        st.markdown('<div class="info-header">업종명</div>', unsafe_allow_html=True)
    with col2:
        st.text_input("업종명", label_visibility="collapsed", key="industry_name")
    with col3:
        st.markdown('<div class="info-header">생산품</div>', unsafe_allow_html=True)
    with col4:
        st.text_input("생산품", label_visibility="collapsed", key="product_name")
    
    col1, col2, col3, col4 = st.columns([1, 2, 1, 2])
    with col1:
        st.markdown('<div class="info-header">원(재)료</div>', unsafe_allow_html=True)
    with col2:
        st.text_input("원재료", label_visibility="collapsed", key="raw_material")
    with col3:
        st.markdown('<div class="info-header">근로자</div>', unsafe_allow_html=True)
    with col4:
        st.text_input("근로자", label_visibility="collapsed", key="workers_info")
    
    st.markdown('<hr style="margin: 30px 0;">', unsafe_allow_html=True)
    
    # 공정(작업)순서 테이블
    st.markdown('<h3 style="text-align: center; color: #1f2937;">공정(작업)순서</h3>', unsafe_allow_html=True)
    
    # 원본과 동일한 테이블 헤더
    st.markdown("""
    <table class="process-table">
        <tr>
            <th rowspan="2" class="process-header" style="width: 8%;">공정<br>(작업)순서</th>
            <th colspan="2" class="process-header">기계기구 및 설비명</th>
            <th colspan="3" class="process-header">유해화학물질</th>
            <th colspan="8" class="process-header">기타 안전보건상 정보</th>
        </tr>
        <tr>
            <th class="sub-header">기계기구 및<br>설비명</th>
            <th class="sub-header">수량</th>
            <th class="sub-header">화학물질명</th>
            <th class="sub-header">취급량/일</th>
            <th class="sub-header">취급시간</th>
            <th class="sub-header">3년간<br>재해사례</th>
            <th class="sub-header">앗차<br>사고사례</th>
            <th class="sub-header">근로자<br>구성및특성</th>
            <th class="sub-header">도급/교대<br>작업유무</th>
            <th class="sub-header">운반수단</th>
            <th class="sub-header">안전작업<br>허가증<br>필요작업</th>
            <th class="sub-header">작업환경<br>측정유무</th>
            <th class="sub-header">특별안전<br>교육대상</th>
        </tr>
    </table>
    """, unsafe_allow_html=True)
    
    # 데이터 저장을 위한 리스트
    process_data_list = []
    
    # 공정별 데이터 입력
    if 'processes' in st.session_state:
        for idx, process in enumerate(st.session_state.processes):
            if process['name']:
                # 균등한 컬럼 분할
                cols = st.columns([0.8, 1.2, 0.5, 1.2, 0.6, 0.6, 0.8, 0.8, 0.8, 0.8, 0.8, 0.8, 0.8, 0.8])
                
                # 각 필드의 값을 저장할 딕셔너리
                process_row = {}
                
                # 공정(작업)순서
                with cols[0]:
                    process_row['공정순서'] = process['name']
                    st.text_input(f"공정_{idx}", value=process['name'], disabled=True, label_visibility="collapsed")
                
                # 기계기구 및 설비명
                with cols[1]:
                    process_row['기계기구및설비명'] = process['equipment']
                    st.text_area(f"기계_{idx}", value=process['equipment'], height=100, disabled=True, label_visibility="collapsed")
                
                # 수량
                with cols[2]:
                    qty = st.text_input(f"수량_{idx}", placeholder="", label_visibility="collapsed", key=f"qty_{idx}")
                    process_row['수량'] = qty
                
                # 화학물질명
                with cols[3]:
                    process_row['화학물질명'] = process['hazardous_material']
                    st.text_area(f"화학_{idx}", value=process['hazardous_material'], height=100, disabled=True, label_visibility="collapsed")
                
                # 취급량/일
                with cols[4]:
                    amount = st.text_input(f"취급량_{idx}", placeholder="", label_visibility="collapsed", key=f"amount_{idx}")
                    process_row['취급량/일'] = amount
                
                # 취급시간
                with cols[5]:
                    time = st.text_input(f"취급시간_{idx}", placeholder="", label_visibility="collapsed", key=f"time_{idx}")
                    process_row['취급시간'] = time
                
                # 3년간 재해사례
                with cols[6]:
                    accident = st.text_input(f"재해사례_{idx}", placeholder="", label_visibility="collapsed", key=f"accident_{idx}")
                    process_row['3년간재해사례'] = accident
                
                # 앗차사고사례
                with cols[7]:
                    near_miss = st.text_input(f"앗차_{idx}", placeholder="", label_visibility="collapsed", key=f"near_miss_{idx}")
                    process_row['앗차사고사례'] = near_miss
                
                # 근로자 구성및특성
                with cols[8]:
                    workers = st.text_input(f"근로자구성_{idx}", placeholder="", label_visibility="collapsed", key=f"workers_{idx}")
                    process_row['근로자구성및특성'] = workers
                
                # 도급/교대 작업유무
                with cols[9]:
                    contract = st.selectbox(f"도급_{idx}", ["", "유", "무"], label_visibility="collapsed", key=f"contract_{idx}")
                    process_row['도급/교대작업유무'] = contract
                
                # 운반수단
                with cols[10]:
                    transport = st.text_input(f"운반_{idx}", placeholder="", label_visibility="collapsed", key=f"transport_{idx}")
                    process_row['운반수단'] = transport
                
                # 안전작업허가증필요작업
                with cols[11]:
                    permit = st.selectbox(f"허가증_{idx}", ["", "유", "무"], label_visibility="collapsed", key=f"permit_{idx}")
                    process_row['안전작업허가증필요작업'] = permit
                
                # 작업환경측정유무
                with cols[12]:
                    measurement = st.selectbox(f"측정_{idx}", ["", "유", "무"], label_visibility="collapsed", key=f"measurement_{idx}")
                    process_row['작업환경측정유무'] = measurement
                
                # 특별안전교육대상
                with cols[13]:
                    special_edu = st.text_input(f"특별교육_{idx}", placeholder="", label_visibility="collapsed", key=f"special_edu_{idx}")
                    process_row['특별안전교육대상'] = special_edu
                
                process_data_list.append(process_row)
                st.markdown('<hr style="margin: 10px 0; border: 0; border-top: 1px solid #d97706;">', unsafe_allow_html=True)
    
    # 데이터 저장 버튼 (엑셀)
    st.markdown('<br>', unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 1, 1])
    with col2:
        if st.button("💾 위험정보 엑셀 저장", use_container_width=True, key="save_tab3"):
            output = BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # 빈 데이터프레임으로 시트 생성
                df = pd.DataFrame()
                df.to_excel(writer, sheet_name='위험정보', index=False)
                
                workbook = writer.book
                worksheet = writer.sheets['위험정보']
                
                # 서식 설정
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
                
                header_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                # 상단 정보 테이블
                current_row = 1
                
                # 업종명, 생산품
                worksheet['A1'].value = "업종명"
                worksheet['A1'].fill = header_fill
                worksheet['A1'].alignment = center_align
                worksheet['A1'].border = thin_border
                
                worksheet.merge_cells('B1:C1')
                worksheet['B1'].value = st.session_state.get('industry_name', '')
                worksheet['B1'].alignment = center_align
                worksheet['B1'].border = thin_border
                
                worksheet['D1'].value = "생산품"
                worksheet['D1'].fill = header_fill
                worksheet['D1'].alignment = center_align
                worksheet['D1'].border = thin_border
                
                worksheet.merge_cells('E1:F1')
                worksheet['E1'].value = st.session_state.get('product_name', '')
                worksheet['E1'].alignment = center_align
                worksheet['E1'].border = thin_border
                
                # 원(재)료, 근로자
                worksheet['A2'].value = "원(재)료"
                worksheet['A2'].fill = header_fill
                worksheet['A2'].alignment = center_align
                worksheet['A2'].border = thin_border
                
                worksheet.merge_cells('B2:C2')
                worksheet['B2'].value = st.session_state.get('raw_material', '')
                worksheet['B2'].alignment = center_align
                worksheet['B2'].border = thin_border
                
                worksheet['D2'].value = "근로자"
                worksheet['D2'].fill = header_fill
                worksheet['D2'].alignment = center_align
                worksheet['D2'].border = thin_border
                
                worksheet.merge_cells('E2:F2')
                worksheet['E2'].value = st.session_state.get('workers_info', '')
                worksheet['E2'].alignment = center_align
                worksheet['E2'].border = thin_border
                
                # 공정(작업)순서 테이블 헤더
                current_row = 4
                
                # 첫 번째 행
                worksheet.merge_cells(f'A{current_row}:A{current_row+1}')
                worksheet[f'A{current_row}'].value = "공정\n(작업)순서"
                worksheet[f'A{current_row}'].fill = header_fill
                worksheet[f'A{current_row}'].alignment = center_align
                worksheet[f'A{current_row}'].border = thin_border
                
                worksheet.merge_cells(f'B{current_row}:C{current_row}')
                worksheet[f'B{current_row}'].value = "기계기구 및 설비명"
                worksheet[f'B{current_row}'].fill = header_fill
                worksheet[f'B{current_row}'].alignment = center_align
                worksheet[f'B{current_row}'].border = thin_border
                
                worksheet.merge_cells(f'D{current_row}:F{current_row}')
                worksheet[f'D{current_row}'].value = "유해화학물질"
                worksheet[f'D{current_row}'].fill = header_fill
                worksheet[f'D{current_row}'].alignment = center_align
                worksheet[f'D{current_row}'].border = thin_border
                
                worksheet.merge_cells(f'G{current_row}:N{current_row}')
                worksheet[f'G{current_row}'].value = "기타 안전보건상 정보"
                worksheet[f'G{current_row}'].fill = header_fill
                worksheet[f'G{current_row}'].alignment = center_align
                worksheet[f'G{current_row}'].border = thin_border
                
                # 두 번째 행
                current_row += 1
                headers = ['기계기구 및\n설비명', '수량', '화학물질명', '취급량/일', '취급시간',
                          '3년간\n재해사례', '앗차\n사고사례', '근로자\n구성및특성', '도급/교대\n작업유무',
                          '운반수단', '안전작업\n허가증\n필요작업', '작업환경\n측정유무', '특별안전\n교육대상']
                
                col_idx = 1  # B열부터 시작
                for header in headers:
                    cell = worksheet[f'{get_column_letter(col_idx+1)}{current_row}']
                    cell.value = header
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = thin_border
                    col_idx += 1
                
                # 데이터 입력
                current_row += 1
                
                if 'processes' in st.session_state:
                    for idx, process in enumerate(st.session_state.processes):
                        if process['name']:
                            # 공정(작업)순서
                            worksheet[f'A{current_row}'].value = process['name']
                            worksheet[f'A{current_row}'].alignment = center_align
                            worksheet[f'A{current_row}'].border = thin_border
                            
                            # 기계기구 및 설비명
                            worksheet[f'B{current_row}'].value = process['equipment']
                            worksheet[f'B{current_row}'].alignment = left_align
                            worksheet[f'B{current_row}'].border = thin_border
                            
                            # 수량
                            worksheet[f'C{current_row}'].value = st.session_state.get(f'qty_{idx}', '')
                            worksheet[f'C{current_row}'].alignment = center_align
                            worksheet[f'C{current_row}'].border = thin_border
                            
                            # 화학물질명
                            worksheet[f'D{current_row}'].value = process['hazardous_material']
                            worksheet[f'D{current_row}'].alignment = left_align
                            worksheet[f'D{current_row}'].border = thin_border
                            
                            # 취급량/일
                            worksheet[f'E{current_row}'].value = st.session_state.get(f'amount_{idx}', '')
                            worksheet[f'E{current_row}'].alignment = center_align
                            worksheet[f'E{current_row}'].border = thin_border
                            
                            # 취급시간
                            worksheet[f'F{current_row}'].value = st.session_state.get(f'time_{idx}', '')
                            worksheet[f'F{current_row}'].alignment = center_align
                            worksheet[f'F{current_row}'].border = thin_border
                            
                            # 3년간 재해사례
                            worksheet[f'G{current_row}'].value = st.session_state.get(f'accident_{idx}', '')
                            worksheet[f'G{current_row}'].alignment = center_align
                            worksheet[f'G{current_row}'].border = thin_border
                            
                            # 앗차사고사례
                            worksheet[f'H{current_row}'].value = st.session_state.get(f'near_miss_{idx}', '')
                            worksheet[f'H{current_row}'].alignment = center_align
                            worksheet[f'H{current_row}'].border = thin_border
                            
                            # 근로자 구성및특성
                            worksheet[f'I{current_row}'].value = st.session_state.get(f'workers_{idx}', '')
                            worksheet[f'I{current_row}'].alignment = center_align
                            worksheet[f'I{current_row}'].border = thin_border
                            
                            # 도급/교대 작업유무
                            worksheet[f'J{current_row}'].value = st.session_state.get(f'contract_{idx}', '')
                            worksheet[f'J{current_row}'].alignment = center_align
                            worksheet[f'J{current_row}'].border = thin_border
                            
                            # 운반수단
                            worksheet[f'K{current_row}'].value = st.session_state.get(f'transport_{idx}', '')
                            worksheet[f'K{current_row}'].alignment = center_align
                            worksheet[f'K{current_row}'].border = thin_border
                            
                            # 안전작업허가증필요작업
                            worksheet[f'L{current_row}'].value = st.session_state.get(f'permit_{idx}', '')
                            worksheet[f'L{current_row}'].alignment = center_align
                            worksheet[f'L{current_row}'].border = thin_border
                            
                            # 작업환경측정유무
                            worksheet[f'M{current_row}'].value = st.session_state.get(f'measurement_{idx}', '')
                            worksheet[f'M{current_row}'].alignment = center_align
                            worksheet[f'M{current_row}'].border = thin_border
                            
                            # 특별안전교육대상
                            worksheet[f'N{current_row}'].value = st.session_state.get(f'special_edu_{idx}', '')
                            worksheet[f'N{current_row}'].alignment = center_align
                            worksheet[f'N{current_row}'].border = thin_border
                            
                            current_row += 1
                
                # 열 너비 조정
                worksheet.column_dimensions['A'].width = 12
                worksheet.column_dimensions['B'].width = 20
                worksheet.column_dimensions['C'].width = 8
                worksheet.column_dimensions['D'].width = 20
                worksheet.column_dimensions['E'].width = 10
                worksheet.column_dimensions['F'].width = 10
                worksheet.column_dimensions['G'].width = 10
                worksheet.column_dimensions['H'].width = 10
                worksheet.column_dimensions['I'].width = 12
                worksheet.column_dimensions['J'].width = 10
                worksheet.column_dimensions['K'].width = 10
                worksheet.column_dimensions['L'].width = 12
                worksheet.column_dimensions['M'].width = 12
                worksheet.column_dimensions['N'].width = 12
            
            # 다운로드 링크 생성
            output.seek(0)
            b64 = base64.b64encode(output.read()).decode()
            href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="위험성평가_위험정보_{datetime.now().strftime("%Y%m%d")}.xlsx">📥 엑셀 파일 다운로드</a>'
            st.markdown(href, unsafe_allow_html=True)
            st.success("위험정보가 엑셀 파일로 저장되었습니다!")

with tab4:
    st.markdown('<h2 style="text-align: center; color: #1f2937;">유해위험요인 분류</h2>', unsafe_allow_html=True)
    
    # 유해위험요인 분류 기준 데이터
    hazard_categories = {
        '기계(설비)적 요인': [
            [('1.1 끼임(협착/감김/말림)', False), ('1.2 위험한 표면(절단·베임·찔림)', False), ('1.3 기계(설비)의 맞음, 터짐, 끼임, 뒤덮힘, 넘어짐/깔림 위험부분', False)],
            [('1.4 부딪힘 위험 부분', False), ('1.5 넘어짐(미끄러짐·걸림·헛디딤)', False), ('1.6 떨어짐 위험 부분(개구부 등)', False)],
        ],
        '전기적 요인': [
            [('2.1 감전(누전현상 포함)', False), ('2.2 아크', False), ('2.3 정전기', False)],
            [('2.4 화재/폭발 위험', False), ('', False), ('', False)],
        ],
        '화학(물질)적 요인': [
            [('3.1 가스', False), ('3.2 증기', False), ('3.3 에어로졸·흄', False)],
            [('3.4 액체·미스트', False), ('3.5 고체(분진)', False), ('3.6 반응성 물질', False)],
            [('3.7 방사선', False), ('3.8 화재·폭발위험', False), ('3.9 복사열·폭발과압', False)],
        ],
        '생물학적 요인': [
            [('4.1 병원성 미생물,바이러스에 의한 감염', False), ('4.2 유전자 변형물질(GMO)', False), ('4.3 알러지 및 미생물', False)],
            [('4.4 동물', False), ('4.5 식물', False), ('', False)],
        ],
        '작업특성 요인': [
            [('5.1 소음', False), ('5.2 초음파·초저주파음', False), ('5.3 진동', False)],
            [('5.4 근로자 실수(휴먼에러)', False), ('5.5 저압 또는 고압상태', False), ('5.6 질식위험·산소결핍', False)],
            [('5.7 중량물취급작업', False), ('5.8 반복작업', False), ('5.9 불안정한 작업자세', False)],
            [('5.10 작업(조작) 도구', False), ('5.11 기후 / 고온 / 한랭', False), ('', False)],
        ],
        '작업환경 요인': [
            [('6.1 기후·고온·한랭', False), ('6.2 조명', False), ('6.3 공간 및 이동통로', False)],
            [('6.4 주변 근로자', False), ('6.5 작업시간', False), ('6.6 조직 안전문화', False)],
            [('6.7 화상', False), ('', False), ('', False)],
        ]
    }
    
    # 세션 상태 초기화
    if 'hazard_classifications' not in st.session_state:
        st.session_state.hazard_classifications = {}
    
    # 공정별로 유해위험요인 분류표 생성
    if 'processes' in st.session_state:
        for idx, process in enumerate(st.session_state.processes):
            if process['name']:  # 공정명이 있는 경우만
                # 각 공정별 데이터 저장을 위한 키
                process_key = f"hazard_{idx}"
                if process_key not in st.session_state.hazard_classifications:
                    st.session_state.hazard_classifications[process_key] = {
                        'manufacturing_process': '',
                        'classification_code': ''
                    }
                
                # 상단 테이블 생성
                st.markdown(f"""
                <style>
                    .hazard-header {{
                        width: 100%;
                        border-collapse: collapse;
                        margin-bottom: 0px;
                    }}
                    .hazard-header td {{
                        border: 1px solid #000;
                        padding: 10px;
                        text-align: center;
                    }}
                    .header-cell {{
                        background-color: #fef3c7;
                        font-weight: bold;
                    }}
                    .input-cell {{
                        background-color: white;
                        height: 40px;
                    }}
                </style>
                
                <table class="hazard-header">
                    <tr>
                        <td rowspan="2" class="header-cell" style="width: 10%;">제조 공정</td>
                        <td colspan="2" class="header-cell" style="width: 40%; font-size: 18px;">유해위험요인 분류</td>
                        <td rowspan="2" class="header-cell" style="width: 25%;">세부 공정</td>
                        <td rowspan="2" class="header-cell" style="width: 25%;">분류 코드</td>
                    </tr>
                    <tr>
                        <td class="input-cell" colspan="2"></td>
                    </tr>
                </table>
                """, unsafe_allow_html=True)
                
                # 입력 필드를 테이블 위에 오버레이
                col1, col2, col3 = st.columns([1, 2.5, 1.5])
                
                with col1:
                    mfg_process = st.text_input(
                        "제조공정", 
                        value=st.session_state.hazard_classifications[process_key]['manufacturing_process'],
                        label_visibility="collapsed", 
                        key=f"mfg_{process_key}"
                    )
                    st.session_state.hazard_classifications[process_key]['manufacturing_process'] = mfg_process
                
                with col2:
                    subcol1, subcol2 = st.columns([1, 1])
                    with subcol2:
                        # 세부공정 - 탭2의 공정명 자동입력
                        st.text_input(
                            "세부공정", 
                            value=process['name'], 
                            disabled=True, 
                            label_visibility="collapsed", 
                            key=f"subprocess_{process_key}"
                        )
                
                with col3:
                    class_code = st.text_input(
                        "분류코드", 
                        value=st.session_state.hazard_classifications[process_key]['classification_code'],
                        label_visibility="collapsed", 
                        key=f"class_{process_key}"
                    )
                    st.session_state.hazard_classifications[process_key]['classification_code'] = class_code
                
                # 유해위험요인 분류 테이블
                st.markdown("""
                <table style="width: 100%; border-collapse: collapse; margin-top: 10px;">
                    <tr>
                        <th style="border: 1px solid #000; background-color: #fef3c7; text-align: center; padding: 10px; width: 5%;">분류</th>
                        <th style="border: 1px solid #000; background-color: #fef3c7; text-align: center; padding: 10px; width: 15%;">분야</th>
                        <th colspan="3" style="border: 1px solid #000; background-color: #fef3c7; text-align: center; padding: 10px;">유해위험요인</th>
                    </tr>
                </table>
                """, unsafe_allow_html=True)
                
                # 각 카테고리별로 행 생성
                for category_idx, (category, items) in enumerate(hazard_categories.items()):
                    row_count = len(items)
                    
                    # 카테고리별 컨테이너
                    with st.container():
                        st.markdown('<div class="hazard-table">', unsafe_allow_html=True)
                        
                        for row_idx, item_list in enumerate(items):
                            cols = st.columns([0.5, 1.5, 2.5, 2.5, 2.5])
                            
                            # 분류 번호 (카테고리당 한 번만)
                            with cols[0]:
                                if row_idx == 0:
                                    st.markdown(f"""
                                    <div style="border: 1px solid #000; background-color: #fef3c7; 
                                               text-align: center; padding: {25 * row_count}px 5px; 
                                               font-weight: bold; height: {50 * row_count}px;
                                               display: flex; align-items: center; justify-content: center;">
                                        {category_idx + 1}
                                    </div>
                                    """, unsafe_allow_html=True)
                            
                            # 분야 (카테고리당 한 번만)
                            with cols[1]:
                                if row_idx == 0:
                                    st.markdown(f"""
                                    <div style="border: 1px solid #000; background-color: #fef3c7; 
                                               text-align: center; padding: {25 * row_count}px 5px; 
                                               font-weight: bold; height: {50 * row_count}px;
                                               display: flex; align-items: center; justify-content: center;">
                                        {category}
                                    </div>
                                    """, unsafe_allow_html=True)
                            
                            # 유해위험요인 체크박스들 (3개 열에 분산)
                            for sub_idx in range(3):
                                with cols[2 + sub_idx]:
                                    if sub_idx < len(item_list) and item_list[sub_idx][0]:
                                        checked = st.checkbox(
                                            item_list[sub_idx][0], 
                                            key=f"cb_{process_key}_{category_idx}_{row_idx}_{sub_idx}"
                                        )
                                        # 체크박스 상태 저장
                                        if 'checkboxes' not in st.session_state.hazard_classifications[process_key]:
                                            st.session_state.hazard_classifications[process_key]['checkboxes'] = {}
                                        st.session_state.hazard_classifications[process_key]['checkboxes'][f"{category_idx}_{row_idx}_{sub_idx}"] = {
                                            'name': item_list[sub_idx][0],
                                            'checked': checked
                                        }
                                    else:
                                        st.write("")
                        
                        st.markdown('</div>', unsafe_allow_html=True)
                
                st.markdown('<hr style="margin: 30px 0; border: 2px solid #000;">', unsafe_allow_html=True)
    
    # 데이터 저장 버튼
    st.markdown('<br>', unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 1, 1])
    with col2:
        if st.button("💾 유해위험요인 분류 엑셀 저장", use_container_width=True, key="save_tab4"):
            output = BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # 빈 데이터프레임으로 시트 생성
                df = pd.DataFrame()
                df.to_excel(writer, sheet_name='유해위험요인분류', index=False)
                
                workbook = writer.book
                worksheet = writer.sheets['유해위험요인분류']
                
                # 서식 설정
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
                
                header_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                current_row = 1
                
                # 각 공정별로 테이블 생성
                if 'processes' in st.session_state:
                    for idx, process in enumerate(st.session_state.processes):
                        if process['name']:
                            process_key = f"hazard_{idx}"
                            
                            # 테이블 헤더
                            worksheet.merge_cells(f'A{current_row}:B{current_row}')
                            worksheet[f'A{current_row}'].value = "제조 공정"
                            worksheet[f'A{current_row}'].fill = header_fill
                            worksheet[f'A{current_row}'].alignment = center_align
                            worksheet[f'A{current_row}'].border = thin_border
                            
                            worksheet.merge_cells(f'C{current_row}:D{current_row}')
                            worksheet[f'C{current_row}'].value = "유해위험요인 분류"
                            worksheet[f'C{current_row}'].fill = header_fill
                            worksheet[f'C{current_row}'].alignment = center_align
                            worksheet[f'C{current_row}'].border = thin_border
                            worksheet[f'C{current_row}'].font = Font(bold=True, size=14)
                            
                            worksheet[f'E{current_row}'].value = "세부 공정"
                            worksheet[f'E{current_row}'].fill = header_fill
                            worksheet[f'E{current_row}'].alignment = center_align
                            worksheet[f'E{current_row}'].border = thin_border
                            
                            worksheet[f'F{current_row}'].value = "분류 코드"
                            worksheet[f'F{current_row}'].fill = header_fill
                            worksheet[f'F{current_row}'].alignment = center_align
                            worksheet[f'F{current_row}'].border = thin_border
                            
                            current_row += 1
                            
                            # 입력 데이터
                            worksheet.merge_cells(f'A{current_row}:B{current_row}')
                            worksheet[f'A{current_row}'].value = st.session_state.hazard_classifications.get(process_key, {}).get('manufacturing_process', '')
                            worksheet[f'A{current_row}'].alignment = center_align
                            worksheet[f'A{current_row}'].border = thin_border
                            
                            worksheet.merge_cells(f'C{current_row}:D{current_row}')
                            worksheet[f'C{current_row}'].border = thin_border
                            
                            worksheet[f'E{current_row}'].value = process['name']
                            worksheet[f'E{current_row}'].alignment = center_align
                            worksheet[f'E{current_row}'].border = thin_border
                            
                            worksheet[f'F{current_row}'].value = st.session_state.hazard_classifications.get(process_key, {}).get('classification_code', '')
                            worksheet[f'F{current_row}'].alignment = center_align
                            worksheet[f'F{current_row}'].border = thin_border
                            
                            current_row += 2
                            
                            # 유해위험요인 분류 테이블 헤더
                            worksheet[f'A{current_row}'].value = "분류"
                            worksheet[f'A{current_row}'].fill = header_fill
                            worksheet[f'A{current_row}'].alignment = center_align
                            worksheet[f'A{current_row}'].border = thin_border
                            
                            worksheet[f'B{current_row}'].value = "분야"
                            worksheet[f'B{current_row}'].fill = header_fill
                            worksheet[f'B{current_row}'].alignment = center_align
                            worksheet[f'B{current_row}'].border = thin_border
                            
                            worksheet.merge_cells(f'C{current_row}:E{current_row}')
                            worksheet[f'C{current_row}'].value = "유해위험요인"
                            worksheet[f'C{current_row}'].fill = header_fill
                            worksheet[f'C{current_row}'].alignment = center_align
                            worksheet[f'C{current_row}'].border = thin_border
                            
                            current_row += 1
                            
                            # 각 카테고리별 데이터
                            for cat_idx, (category, items) in enumerate(hazard_categories.items()):
                                start_row = current_row
                                
                                # 카테고리 내 모든 항목 출력
                                for row_idx, item_list in enumerate(items):
                                    # 분류 번호 (카테고리의 첫 번째 행에만)
                                    if row_idx == 0:
                                        worksheet[f'A{current_row}'].value = cat_idx + 1
                                        worksheet[f'A{current_row}'].alignment = center_align
                                        worksheet[f'A{current_row}'].border = thin_border
                                        
                                        worksheet[f'B{current_row}'].value = category
                                        worksheet[f'B{current_row}'].alignment = center_align
                                        worksheet[f'B{current_row}'].border = thin_border
                                    else:
                                        worksheet[f'A{current_row}'].border = thin_border
                                        worksheet[f'B{current_row}'].border = thin_border
                                    
                                    # 유해위험요인 항목들
                                    col_letters = ['C', 'D', 'E']
                                    for sub_idx, (item_name, _) in enumerate(item_list):
                                        if item_name:
                                            cell_value = item_name
                                            # 체크박스 상태 확인
                                            checkbox_key = f"{cat_idx}_{row_idx}_{sub_idx}"
                                            if process_key in st.session_state.hazard_classifications:
                                                checkboxes = st.session_state.hazard_classifications[process_key].get('checkboxes', {})
                                                if checkbox_key in checkboxes and checkboxes[checkbox_key]['checked']:
                                                    cell_value = f"☑ {item_name}"
                                                else:
                                                    cell_value = f"☐ {item_name}"
                                            
                                            worksheet[f'{col_letters[sub_idx]}{current_row}'].value = cell_value
                                            worksheet[f'{col_letters[sub_idx]}{current_row}'].alignment = left_align
                                            worksheet[f'{col_letters[sub_idx]}{current_row}'].border = thin_border
                                        else:
                                            worksheet[f'{col_letters[sub_idx]}{current_row}'].border = thin_border
                                    
                                    current_row += 1
                                
                                # 카테고리별로 병합
                                if len(items) > 1:
                                    worksheet.merge_cells(f'A{start_row}:A{current_row-1}')
                                    worksheet.merge_cells(f'B{start_row}:B{current_row-1}')
                            
                            current_row += 2  # 공정 간 간격
                
                # 열 너비 조정
                worksheet.column_dimensions['A'].width = 8
                worksheet.column_dimensions['B'].width = 20
                worksheet.column_dimensions['C'].width = 25
                worksheet.column_dimensions['D'].width = 25
                worksheet.column_dimensions['E'].width = 25
                worksheet.column_dimensions['F'].width = 15
            
            # 다운로드 링크 생성
            output.seek(0)
            b64 = base64.b64encode(output.read()).decode()
            download_filename = f"위험성평가_유해위험요인분류_{datetime.now().strftime('%Y%m%d')}.xlsx"
            href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{download_filename}">📥 엑셀 파일 다운로드</a>'
            st.markdown(href, unsafe_allow_html=True)
            st.success("유해위험요인 분류가 엑셀 파일로 저장되었습니다!")
